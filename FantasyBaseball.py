# -*- coding: utf-8 -*-
"""
Created on Wed Feb 21 21:30:51 2024

@author: homer
"""


# import pandas as pd
# # import openpyxl

# # Read the Excel file
# file_path = r'C:\Users\homer\.spyder-py3\Cheatsheet.xlsx'
# new_file_path = r'C:\Users\homer\.spyder-py3\Modified_Cheatsheet.xlsx'

# # #Set veryHidden sheets to visible
# # # Open the workbook
# # workbook = openpyxl.load_workbook(file_path)
# # # Iterate through all sheets
# # for sheet_name in workbook.sheetnames:
# #     sheet = workbook[sheet_name] 
# #     # Check if the sheet is very hidden
# #     print(sheet.sheet_state)
# #     if sheet.sheet_state == 'veryHidden':
# #         # Unhide the sheet
# #         sheet.sheet_state = 'visible'
# # # Save the changes
# # workbook.save(file_path)
# # # Close the workbook
# # workbook.close()

# #Delete NL
# xls = pd.ExcelFile(file_path)

# # Define the sheets to ignore
# include_sheets = ['Settings', 'My Team']

# # Define the sheets where 'LG' should be checked
# lg_sheets = ['Hitters', '1B', '2B', '3B', 'C', 'OF', 'SS', '1B+3B', '2B+SS', 'UTIL', 'Pitchers', 'SP', 'RP', 'SP+RP']

# # Create a dictionary to store modified DataFrames
# modified_dfs = {}

# # Iterate through sheets
# for sheet_name in xls.sheet_names:
#     print(sheet_name)
    
#     # Read the sheet
#     df = xls.parse(sheet_name)
    
#     # Include sheets as is
#     if sheet_name in include_sheets:
#         modified_dfs[sheet_name] = df
#         continue
    
#     # Check and delete rows based on sheet name
#     if sheet_name == 'Standard Deviations':
#         teams_to_keep = ['ARI', 'ATL', 'CHC', 'CIN', 'COL', 'LAD', 'MIA', 'MIL', 'NYM', 'PHI', 'PIT', 'SD', 'SF', 'STL', 'WSH']
#         df = df[~df['Team'].isin(teams_to_keep)]
#     elif sheet_name == 'Rankings':
#         df = df[df['League'] != 'NL']
#     elif sheet_name in lg_sheets:
#         df = df[df['LG'] != 'NL']
    
#     # Store the modified DataFrame in the dictionary
#     modified_dfs[sheet_name] = df

# # Save the modified sheets to a new Excel file
# with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
#     for sheet_name, df in modified_dfs.items():
#         df.to_excel(writer, sheet_name, index=False)

# # Copy the ignored sheets as is to the new workbook
# with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
#     for sheet_name in include_sheets:
#         df = xls.parse(sheet_name)
#         df.to_excel(writer, sheet_name, index=False)
        
        
        
        
# import openpyxl

# # Update the file path
# file_path = r'C:\Users\homer\.spyder-py3\Cheatsheet.xlsx'

# # Open the workbook
# workbook = openpyxl.load_workbook(file_path)

# # Define the sheets to ignore
# ignore_sheets = ['Settings', 'My Team']

# # Define the sheets where 'LG' should be checked
# lg_sheets = ['Hitters', '1B', '2B', '3B', 'C', 'OF', 'SS', '1B+3B', '2B+SS', 'UTIL', 'Pitchers', 'SP', 'RP', 'SP+RP']

# # Iterate through sheets
# for sheet_name in workbook.sheetnames:
#     if sheet_name not in ignore_sheets:
#         # Get the active sheet
#         sheet = workbook[sheet_name]

#         # Check and delete rows based on sheet name
#         if sheet_name == 'Standard Deviations':
#             teams_to_keep = ['ARI', 'ATL', 'CHC', 'CIN', 'COL', 'LAD', 'MIA', 'MIL', 'NYM', 'PHI', 'PIT', 'SD', 'SF', 'STL', 'WSH']
#             rows_to_delete = []
#             for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
#                 if row[1].value not in teams_to_keep:
#                     rows_to_delete.append(row)
#             for row in rows_to_delete:
#                 sheet.delete_rows(row[0].row)  # Use row[0].row for the row index
        
#         elif sheet_name == 'Rankings':
#             rows_to_delete = []
#             for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
#                 if row[2].value == 'NL':
#                     rows_to_delete.append(row)
#             for row in rows_to_delete:
#                 sheet.delete_rows(row[0].row)
        
#         elif sheet_name in lg_sheets:
#             rows_to_delete = []
#             for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
#                 if row[3].value == 'NL':
#                     rows_to_delete.append(row)
#             for row in rows_to_delete:
#                 sheet.delete_rows(row[0].row)

# # Save the changes
# workbook.save(r'C:\Users\homer\.spyder-py3\Modified_Cheatsheet.xlsx')

# Close the workbook
# workbook.close()

# import openpyxl

# # Update the file paths
# original_file_path = r'C:\Users\homer\.spyder-py3\Cheatsheet.xlsx'
# new_file_path = r'C:\Users\homer\.spyder-py3\Modified_Cheatsheet.xlsx'

# # Define the sheets where 'LG' should be checked
# lg_sheets = ['Hitters', '1B', '2B', '3B', 'C', 'OF', 'SS', '1B+3B', '2B+SS', 'UTIL', 'Pitchers', 'SP', 'RP', 'SP+RP']

# # Open the original workbook
# original_workbook = openpyxl.load_workbook(original_file_path, data_only=False)

# # Create a new workbook
# new_workbook = openpyxl.Workbook()

# # Iterate through sheets
# for sheet_name in original_workbook.sheetnames:
#     original_sheet = original_workbook[sheet_name]
#     new_sheet = new_workbook.create_sheet(title=sheet_name)

#     # Check and delete rows based on sheet name
#     if sheet_name == 'Standard Deviations':
#         teams_to_keep = ['ARI', 'ATL', 'CHC', 'CIN', 'COL', 'LAD', 'MIA', 'MIL', 'NYM', 'PHI', 'PIT', 'SD', 'SF', 'STL', 'WSH']
#         for row in reversed(list(original_sheet.iter_rows(min_row=2, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column))):
#             if row[1].value not in teams_to_keep:
#                 original_sheet.delete_rows(row[0].row)

#     elif sheet_name == 'Rankings':
#         for row in reversed(list(original_sheet.iter_rows(min_row=2, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column))):
#             if row[2].value == 'NL':
#                 original_sheet.delete_rows(row[0].row)

#     elif sheet_name in lg_sheets:
#         for row in reversed(list(original_sheet.iter_rows(min_row=2, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column))):
#             if row[3].value == 'NL':
#                 original_sheet.delete_rows(row[0].row)

#     # Copy cell values to the new sheet
#     for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column):
#         for cell in row:
#             new_sheet[cell.coordinate].value = cell.value
#             new_sheet[cell.coordinate].data_type = cell.data_type  # Preserve data type (e.g., date, time, etc.)

# # Save the new workbook
# new_workbook.save(new_file_path)


import openpyxl

# Update the file paths
original_file_path = r'C:\Users\homer\Downloads\Cheat-Sheet-Generator-0229.xlsx'
new_file_path = r'C:\Users\homer\Downloads\modifiedcheatsheet.xlsx'

# Define the sheets where 'LG' should be checked
lg_sheets = ['Hitters', '1B', '2B', '3B', 'C', 'OF', 'SS', '1B+3B', '2B+SS', 'UTIL', 'Pitchers', 'SP', 'RP', 'SP+RP']


# Open the original workbook
original_workbook = openpyxl.load_workbook(original_file_path, data_only=False)

# Create a new workbook
new_workbook = openpyxl.Workbook()

# Iterate through sheets
for sheet_name in original_workbook.sheetnames:
    original_sheet = original_workbook[sheet_name]
    new_sheet = new_workbook.create_sheet(title=sheet_name)

    # Check and copy rows based on sheet name
    if sheet_name == 'Standard Deviations':
        teams_to_keep = ['ARI', 'ATL', 'CHC', 'CIN', 'COL', 'LAD', 'MIA', 'MIL', 'NYM', 'PHI', 'PIT', 'SD', 'SF', 'STL', 'WSH']
        for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column):
            if row[1].value in teams_to_keep:
                new_sheet.append([cell.value for cell in row])

    elif sheet_name == 'Rankings':
        for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column):
            if row[2].value != 'NL':
                new_sheet.append([cell.value for cell in row])

    elif sheet_name in lg_sheets:
        for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row, min_col=1, max_col=original_sheet.max_column):
            if row[3].value != 'NL':
                new_sheet.append([cell.value for cell in row])

# Save the new workbook
new_workbook.save(new_file_path)


